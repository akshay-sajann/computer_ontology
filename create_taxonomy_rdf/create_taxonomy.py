from rdflib import Graph, URIRef, Literal, Namespace
from rdflib.namespace import SKOS, RDF, RDFS, SDO
import numpy as np
import pandas as pd
import re
from openpyxl import load_workbook

CHEMINF = Namespace("http://semanticscience.org/resource/CHEMINF_")

schema_uri = 'http://data.odeuropa.eu/vocabulary/expert-taxonomy'
schema_concept = URIRef(schema_uri)

taxonomy_map = {}

expert_dataset_train = '../data/train/expert_train.csv'
expert_dataset_test = '../data/test/expert_test.csv'

def parse_sheet(sheet, g, FAMILY_DESCR_COL, SOURCE_BASED_DESCR, schema_concept=None):
    sheet_name = sheet.title
    family_name = re.sub(r'(?i)_cluster', '', sheet_name).replace('_', ' ')
    
    print('******')
    print(family_name, schema_concept)
    codename = family_name.replace(' ', '_').lower()
    family_uri = f"{schema_uri}/{codename}"
    family_concept = URIRef(family_uri)

    g.add((family_concept, RDF.type, SKOS.ConceptScheme if schema_concept is None else SKOS.Concept))
    g.add((family_concept, SKOS.prefLabel, Literal(family_name, 'en')))

     
    if schema_concept is None:
        schema_concept = family_concept
    else:
        g.add((family_concept, SKOS.topConceptOf, schema_concept))

        

    level=[family_concept, None, None]
    is_flat = True
    # is_flat = len(list(df)) < 4

    # full_previous_line = False
    first_line = True
    for row in sheet.iter_rows():
        if first_line:
            first_line=False
            continue
    # for i, row in df.iterrows():
        for j, cell in enumerate(row):
            col = cell.value
            if not col:
                continue
            if  j == FAMILY_DESCR_COL and col != 'Family descriptors':
                g.add((family_concept, SKOS.altLabel, Literal(col, 'en')))
                taxonomy_map[col] = taxonomy_map.get(col, family_concept)
            elif col == 'Source based descriptors':
                break
            elif j>SOURCE_BASED_DESCR+1:
                g.add((level[-1], SKOS.altLabel, Literal(col, 'en')))
                taxonomy_map[col] = taxonomy_map.get(col, level[-1])
            elif j>=SOURCE_BASED_DESCR:
                lv = j-SOURCE_BASED_DESCR

                ancestor = level[lv]
                if j == SOURCE_BASED_DESCR:
                    if cell.font.bold:
                        is_flat = False
                    elif not is_flat:
                        g.add((level[lv+1], SKOS.altLabel, Literal(col, 'en')))
                        taxonomy_map[col] = taxonomy_map.get(col, level[lv+1])
                        continue

                concept = URIRef(str(ancestor)+'/'+re.sub('[, ]+','_',col.lower()))
                g.add((concept, RDF.type, SKOS.Concept))
                g.add((concept, SKOS.prefLabel, Literal(col, 'en')))
                g.add((concept, SKOS.broader, ancestor))
                g.add((concept, SKOS.inScheme, schema_concept))
                taxonomy_map[col] = taxonomy_map.get(col, concept)

                level[lv+1] = concept



# xls = pd.ExcelFile('source.xlsx')
workbook = load_workbook('source.xlsx', rich_text=True)
main_graph = Graph()
quality_graph = Graph()

main_graph.add((schema_concept, RDF.type, SKOS.ConceptScheme))
main_graph.add((schema_concept, RDFS.label, Literal("Expert Taxonomy")))
main_graph.add((schema_concept, SDO.version, Literal("1.0.0")))


for sheet in workbook.worksheets: # xls.sheet_names:
    sheet_name = sheet.title

    if sheet_name == 'Quality':
        parse_sheet(sheet, quality_graph, FAMILY_DESCR_COL = 0, SOURCE_BASED_DESCR = 1, schema_concept=URIRef(schema_uri+'/quality'))
        quality_graph.serialize(destination='expert_taxonomy-quality.ttl')

        
    if not sheet_name.lower().endswith('cluster'):
        continue

    parse_sheet(sheet, main_graph, FAMILY_DESCR_COL = 1, SOURCE_BASED_DESCR = 2, schema_concept=schema_concept)


    main_graph.serialize(destination='expert_taxonomy.ttl')
    
gc = Graph()
gc.bind("cheminf", CHEMINF)  
gc.bind("skos", SKOS)  


def process_dataset(expert_dataset):
    df = pd.read_csv(expert_dataset)
    for i, x in df.iterrows():
        CID = str(x['CID'])
        uri = 'http://rdf.ncbi.nlm.nih.gov/pubchem/compound/CID' + CID
        molecule = URIRef(uri)

        SMILES_DESCR = CHEMINF['000032'] # Isomeric Smiles descriptor
        SMILES_concept = URIRef(f'http://rdf.ncbi.nlm.nih.gov/pubchem/descriptor/CID{CID}_Isomeric_SMILES')
        gc.add((SMILES_concept, RDF.type, SMILES_DESCR))
        gc.add((SMILES_concept, CHEMINF['CHEMINF_000012'], Literal(x['IsomericSMILES']))) # has value
        gc.add((SMILES_concept, CHEMINF['000143'], molecule)) # is descriptor of
        
        for d in eval(x['Descriptors']):
            gc.add((URIRef(schema_uri+'/'+d), CHEMINF['000143'], molecule))

process_dataset(expert_dataset_train)
process_dataset(expert_dataset_test)

gc.serialize(destination='expert_taxonomy-chemical.ttl')